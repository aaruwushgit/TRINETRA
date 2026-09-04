"""Excel logging — the project's actual deliverable.

**`.xlsx` is not an append-friendly format.** A workbook is a zip archive of
XML; adding a row means loading, mutating and rewriting the whole thing. Two
consequences drive this design:

1. Saving after every plate would rewrite the entire workbook thousands of
   times over a long video.
2. Saving only at the end means a crash — or a Colab session timing out —
   loses the entire run.

So every accepted event is appended to a **JSONL write-ahead log** the moment
it happens, and the workbook is materialized from that log. The JSONL append
is a single line to an open file handle, so it is cheap and survives a crash;
the workbook is regenerated whenever it is flushed. If a run dies, `recover()`
rebuilds the spreadsheet from the log with nothing lost.

The other real-world failure is mundane: **the user has the spreadsheet open
in Excel**, which locks the file on Windows and makes the save raise
`PermissionError`. That must not destroy a run, so the writer keeps the data
in the log and reports the problem rather than crashing.
"""

from __future__ import annotations

import json
from collections.abc import Iterable, Iterator, Sequence
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from alpr.data.schema import Region
from alpr.dedup import DEFAULT_COOLDOWN, Deduplicator

SHEET_NAME = "plates"

# (attribute, header, column width)
COLUMNS: tuple[tuple[str, str, int], ...] = (
    ("timestamp", "Timestamp", 20),
    ("plate", "Plate", 16),
    ("display", "Formatted", 16),
    ("region", "Region", 8),
    ("confidence", "Confidence", 11),
    ("edits", "OCR fixes", 10),
    ("track_id", "Track", 8),
    ("frame", "Frame", 9),
    ("source", "Source", 28),
    ("crop_path", "Crop", 30),
)

_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")


@dataclass(frozen=True)
class PlateEvent:
    """One logged plate sighting."""

    plate: str
    timestamp: datetime
    region: Region = Region.UNKNOWN
    confidence: float = 1.0
    display: str = ""
    edits: int = 0
    track_id: int | None = None
    frame: int | None = None
    source: str | None = None
    crop_path: str | None = None
    extra: dict[str, str] = field(default_factory=dict)

    def to_json(self) -> str:
        payload = asdict(self)
        payload["timestamp"] = self.timestamp.isoformat()
        payload["region"] = self.region.value
        return json.dumps(payload, ensure_ascii=False, sort_keys=True)

    @classmethod
    def from_json(cls, line: str) -> PlateEvent:
        payload = json.loads(line)
        payload["timestamp"] = datetime.fromisoformat(payload["timestamp"])
        payload["region"] = Region(payload.get("region", Region.UNKNOWN.value))
        return cls(**payload)

    def row(self) -> list[object]:
        values: list[object] = []
        for attribute, _, _ in COLUMNS:
            value = getattr(self, attribute)
            if isinstance(value, Region):
                value = value.value
            elif isinstance(value, datetime):
                value = _excel_datetime(value)
            values.append(value)
        return values


def _excel_datetime(when: datetime) -> datetime:
    """Make a datetime writable to a cell.

    openpyxl refuses timezone-aware datetimes outright — Excel has no concept
    of an offset — so an aware value is converted to local time and stripped.
    Dropping the offset silently would be worse: the same run logged from two
    machines would produce times that cannot be compared.
    """
    if when.tzinfo is None:
        return when
    return when.astimezone().replace(tzinfo=None)


class ExcelLogError(RuntimeError):
    """Raised when the workbook cannot be written."""


class ExcelLog:
    """Deduplicating, crash-safe Excel logger.

    Use as a context manager so the workbook is written on the way out::

        with ExcelLog("plates.xlsx") as log:
            log.add(event)
    """

    def __init__(
        self,
        path: str | Path,
        *,
        cooldown=DEFAULT_COOLDOWN,
        deduplicate: bool = True,
        flush_every: int = 50,
    ) -> None:
        self.path = Path(path)
        self.journal_path = self.path.with_suffix(self.path.suffix + ".jsonl")
        self.flush_every = max(1, flush_every)
        self.deduplicator = Deduplicator(cooldown) if deduplicate else None

        self._events: list[PlateEvent] = []
        self._since_flush = 0
        self._journal = None
        self.written = 0

        self.path.parent.mkdir(parents=True, exist_ok=True)

    # -- lifecycle ---------------------------------------------------------

    def __enter__(self) -> ExcelLog:
        self.open()
        return self

    def __exit__(self, *exc_info) -> None:
        self.close()

    def open(self) -> None:
        """Open the journal, recovering any events from an interrupted run."""
        if self.journal_path.exists():
            self._events = list(read_journal(self.journal_path))
            # Re-arm the deduplicator so a resumed run does not re-log a plate
            # that the interrupted one already recorded.
            if self.deduplicator is not None:
                for event in self._events:
                    self.deduplicator.accept(event.plate, event.timestamp)

        self._journal = self.journal_path.open("a", encoding="utf-8")

    def close(self) -> None:
        try:
            self.flush()
        finally:
            if self._journal is not None:
                self._journal.close()
                self._journal = None

    # -- writing -----------------------------------------------------------

    def add(self, event: PlateEvent) -> bool:
        """Record a sighting.

        Returns:
            True if it was logged, False if the deduplicator suppressed it.
        """
        if self._journal is None:
            raise ExcelLogError("ExcelLog is not open — use it as a context manager")

        if self.deduplicator is not None and not self.deduplicator.accept(
            event.plate, event.timestamp
        ):
            return False

        # Journal first, then buffer: an event that reached the log survives a
        # crash even if the workbook is never written.
        self._journal.write(event.to_json() + "\n")
        self._journal.flush()

        self._events.append(event)
        self._since_flush += 1
        if self._since_flush >= self.flush_every:
            self.flush()
        return True

    def flush(self) -> None:
        """Write the workbook from the journalled events."""
        if not self._events:
            return
        write_workbook(self.path, self._events)
        self.written = len(self._events)
        self._since_flush = 0

    @property
    def suppressed(self) -> int:
        return self.deduplicator.suppressed_count if self.deduplicator else 0

    @property
    def logged(self) -> int:
        """Events accepted so far, flushed or not."""
        return len(self._events)

    def summary(self) -> str:
        """One line describing the run.

        Reports events *logged*, not rows flushed. Reporting the flushed count
        made a healthy run read as "0 row(s) written" whenever it was called
        before the first flush — technically true of the workbook, and
        actively misleading about the run, since every event was already
        durable in the journal.
        """
        parts = [f"{self.logged} plate(s) logged to {self.path}"]
        if self.suppressed:
            parts.append(f"{self.suppressed} duplicate(s) suppressed")
        pending = self.logged - self.written
        if pending > 0:
            parts.append(f"{pending} not yet flushed (in {self.journal_path.name})")
        return ", ".join(parts)


def write_workbook(path: str | Path, events: Sequence[PlateEvent]) -> Path:
    """Write events to `path` atomically.

    Builds a temporary file and renames it, so an interrupted write cannot
    leave a truncated workbook where a valid one used to be.

    Raises:
        ExcelLogError: when the destination cannot be replaced — most often
            because the workbook is open in Excel, which locks it on Windows.
    """
    path = Path(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    sheet.append([header for _, header, _ in COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for event in events:
        sheet.append(event.row())

    for index, (attribute, _, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = width
        if attribute == "timestamp":
            for cell in sheet[letter][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        elif attribute == "confidence":
            for cell in sheet[letter][1:]:
                cell.number_format = "0.00"

    # Keep the header visible and filterable — these logs get long.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{sheet.max_row}"

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        workbook.save(tmp)
        tmp.replace(path)
    except PermissionError as exc:
        tmp.unlink(missing_ok=True)
        raise ExcelLogError(
            f"cannot write {path}: it is open in another program. Close it and "
            "flush again — no data is lost, every event is in "
            f"{path.with_suffix(path.suffix + '.jsonl')}."
        ) from exc
    except OSError as exc:
        tmp.unlink(missing_ok=True)
        raise ExcelLogError(f"cannot write {path}: {exc}") from exc

    return path


def read_journal(path: str | Path) -> Iterator[PlateEvent]:
    """Stream events from a write-ahead log, skipping a torn final line.

    A crash mid-write can leave the last line incomplete. That line is the
    only one that may be dropped; everything before it is intact, which is the
    point of journalling one event per line.
    """
    path = Path(path)
    if not path.exists():
        return

    with path.open(encoding="utf-8") as handle:
        lines = handle.readlines()

    for index, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
        try:
            yield PlateEvent.from_json(line)
        except (json.JSONDecodeError, KeyError, TypeError, ValueError):
            if index == len(lines) - 1:
                continue  # torn tail from an interrupted write
            raise


def recover(path: str | Path) -> Path:
    """Rebuild a workbook from its journal after an interrupted run."""
    path = Path(path)
    journal = path.with_suffix(path.suffix + ".jsonl")
    events = list(read_journal(journal))
    if not events:
        raise ExcelLogError(f"nothing to recover: {journal} is empty or missing")
    return write_workbook(path, events)


def read_workbook(path: str | Path) -> list[dict[str, object]]:
    """Read a written log back as dictionaries. Used by tests and reporting."""
    workbook = load_workbook(Path(path), read_only=True)
    sheet = workbook[SHEET_NAME]
    rows: Iterable[tuple] = sheet.iter_rows(values_only=True)
    headers = [header for _, header, _ in COLUMNS]
    out = []
    for index, row in enumerate(rows):
        if index == 0:
            continue
        out.append(dict(zip(headers, row, strict=False)))
    workbook.close()
    return out
